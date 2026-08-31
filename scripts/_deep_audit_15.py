"""Deep audit: Excel 1.1-1.5 vs Strapi vs hide filter vs menu visibility."""
import json
import re
import urllib.request
from pathlib import Path
from collections import defaultdict

import openpyxl

EXCEL = Path(r"d:\riteshglobalca\Copy of Ritesh Arora and Associates_Web Dev Task Breakdown.xlsx")
HIDE = json.loads((Path(__file__).parent.parent / "src/lib/hide-docs-123.json").read_text())
BASE = "https://riteshglobalca.com/strapi/api"

hide_cat = set(HIDE["categories"])
hide_sub = set(HIDE["subcategories"])
hide_svc = set(HIDE["services"])


def fetch_all(endpoint: str):
    page = 1
    rows = []
    while True:
        sep = "&" if "?" in endpoint else "?"
        url = f"{BASE}{endpoint}{sep}pagination[page]={page}&pagination[pageSize]=100"
        with urllib.request.urlopen(url, timeout=60) as r:
            data = json.loads(r.read())
        batch = data.get("data") or []
        rows.extend(batch)
        pag = data.get("meta", {}).get("pagination", {})
        if page >= pag.get("pageCount", 1):
            break
        page += 1
    return rows


def parse_excel_11_to_15():
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    ws = wb["Content"]
    tree = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))  # 1.x -> 1.x.x -> 1.x.x.x pages -> leaves

    current_1x = None  # 1.1, 1.2 ...
    current_subdiv1 = None
    current_subdiv2 = None
    current_page = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row[:6]]
        if not any(cells):
            continue

        # MAIN MODULE column
        if cells[0] and "INDIA" in cells[0].upper():
            continue

        # Detect numbered hierarchy
        for i, cell in enumerate(cells):
            if not cell:
                continue
            m = re.match(r"^(1\.[1-5])(?:\.(\d+))?(?:\.(\d+))?(?:\.(\d+))?(?:\.(\d+))?\s*(.*)$", cell)
            if not m:
                continue
            parts = [p for p in m.groups()[:5] if p]
            label = (m.group(6) or "").strip() or cell
            depth = len(parts)

            if depth == 1 and parts[0] in ("1.1", "1.2", "1.3", "1.4", "1.5"):
                current_1x = parts[0]
                current_subdiv1 = label if label else cell
                tree[current_1x]["__label__"] = current_subdiv1
                current_subdiv2 = None
                current_page = None
            elif depth == 2 and current_1x:
                current_subdiv2 = label if label else cell
                tree[current_1x][current_subdiv2]["__label__"] = current_subdiv2
                current_page = None
            elif depth == 3 and current_1x and current_subdiv2:
                current_page = label if label else cell
                tree[current_1x][current_subdiv2][current_page] = []
            elif depth >= 4 and current_1x and current_subdiv2 and current_page:
                tree[current_1x][current_subdiv2][current_page].append(label if label else cell)
            break

        # Also handle SUB MODULE row style (1.3 Audit appears without 1.3 prefix in col0)
        for cell in cells:
            if re.match(r"^1\.[3-5]\s", cell):
                m = re.match(r"^(1\.[3-5])\s+(.*)", cell)
                if m:
                    current_1x = m.group(1)
                    tree[current_1x]["__label__"] = m.group(2)

    wb.close()
    return tree


# Simpler Excel parse: scan all cells for 1.1-1.5 section headers and page rows
def parse_excel_simple():
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    ws = wb["Content"]

    sections = {f"1.{i}": {"title": "", "subdiv2": {}, "pages": []} for i in range(1, 6)}
    current_section = None
    current_subdiv2 = None

    section_pat = re.compile(r"^(1\.[1-5])\s+(.+)$")
    subdiv2_pat = re.compile(r"^(1\.[1-5]\.\d+)\s+(.+)$")
    page_pat = re.compile(r"^(1\.[1-5]\.\d+\.\d+)\s+(.+)$")
    leaf_pat = re.compile(r"^(1\.[1-5]\.\d+\.\d+\.\d+)\s+(.+)$")

    all_page_codes = set()
    all_subdiv2_codes = set()
    all_section_codes = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        for cell in row:
            if cell is None:
                continue
            text = str(cell).strip()
            if not text:
                continue

            m = section_pat.match(text)
            if m and m.group(1) in sections:
                current_section = m.group(1)
                sections[current_section]["title"] = m.group(2)
                all_section_codes.add(current_section)
                current_subdiv2 = None
                continue

            m = subdiv2_pat.match(text)
            if m:
                code = m.group(1)
                sec = ".".join(code.split(".")[:2])
                if sec in sections:
                    current_section = sec
                    current_subdiv2 = code
                    sections[sec]["subdiv2"][code] = m.group(2)
                    all_subdiv2_codes.add(code)
                continue

            m = page_pat.match(text)
            if m:
                code = m.group(1)
                sec = ".".join(code.split(".")[:2])
                if sec in sections:
                    current_section = sec
                    sections[sec]["pages"].append({"code": code, "title": m.group(2)})
                    all_page_codes.add(code)
                continue

            m = leaf_pat.match(text)
            if m:
                all_page_codes.add(".".join(m.group(1).split(".")[:4]))  # parent page

    wb.close()
    return sections, all_section_codes, all_subdiv2_codes, all_page_codes


sections_excel, sec_codes, subdiv2_codes, page_codes = parse_excel_simple()

mods = fetch_all("/main-modules?fields[0]=title&fields[1]=slug")
cats = fetch_all(
    "/service-categories?fields[0]=title&fields[1]=slug&fields[2]=menuLabel"
    "&populate[mainModule][fields][0]=slug"
)
subs = fetch_all(
    "/service-subcategories?fields[0]=title&fields[1]=slug&fields[2]=menuLabel"
    "&populate[category][fields][0]=slug&populate[category][fields][1]=title"
    "&populate[category][populate][mainModule][fields][0]=slug"
)
services = fetch_all(
    "/services?fields[0]=title&fields[1]=slug"
    "&populate[subcategory][fields][0]=slug&populate[subcategory][fields][1]=title"
    "&populate[subcategory][populate][category][fields][0]=slug"
)

india = next((m for m in mods if m.get("slug") == "india-practice"), None)
india_cats = [c for c in cats if (c.get("mainModule") or {}).get("slug") == "india-practice"]

# Map excel 1.x to expected category slug patterns (from Strapi titles)
EXCEL_TO_CAT_SLUG = {
    "1.1": "gst-indirect-tax-services",
    "1.2": "direct-tax-services",
    "1.3": "audit-assurance-services",
    "1.4": "accounting-compliance-services",
    "1.5": "corporate-advisory-services",
}

print("=" * 80)
print("DEEP AUDIT: EXCEL 1.1–1.5 vs STRAPI vs FRONTEND MENU")
print("=" * 80)

for sec in ["1.1", "1.2", "1.3", "1.4", "1.5"]:
    print(f"\n{'#' * 80}")
    print(f"## EXCEL SECTION {sec}: {sections_excel[sec]['title'] or '(see subdiv rows)'}")
    print(f"{'#' * 80}")

    expected_cat_slug = EXCEL_TO_CAT_SLUG.get(sec)
    strapi_cat = next((c for c in india_cats if c.get("slug") == expected_cat_slug), None)

    subdiv2_in_excel = {k: v for k, v in sections_excel[sec]["subdiv2"].items()}
    pages_in_excel = sections_excel[sec]["pages"]

    print(f"\nExcel subdiv2 count (1.x.y): {len(subdiv2_in_excel)}")
    print(f"Excel page count (1.x.y.z): {len(pages_in_excel)}")

    if not strapi_cat:
        print(f"STRAPI CATEGORY: MISSING — expected slug `{expected_cat_slug}`")
        continue

    cat_hidden = strapi_cat["slug"] in hide_cat
    cat_subs = [s for s in subs if (s.get("category") or {}).get("slug") == strapi_cat["slug"]]
    visible_subs = [s for s in cat_subs if s["slug"] not in hide_sub]
    cat_svcs = [
        s
        for s in services
        if (s.get("subcategory") or {}).get("category", {}).get("slug") == strapi_cat["slug"]
        or any(
            (s.get("subcategory") or {}).get("slug") == sub["slug"]
            for sub in cat_subs
        )
    ]
    # fix service filter
    cat_sub_slugs = {s["slug"] for s in cat_subs}
    cat_svcs = [s for s in services if (s.get("subcategory") or {}).get("slug") in cat_sub_slugs]
    visible_svcs = [s for s in cat_svcs if s["slug"] not in hide_svc and (s.get("subcategory") or {}).get("slug") not in hide_sub]

    print(f"\nSTRAPI CATEGORY: {strapi_cat['title']} (`{strapi_cat['slug']}`)")
    print(f"  Hidden from menu: {'YES — ENTIRE BRANCH SUPPRESSED' if cat_hidden else 'No'}")
    print(f"  Subcategories in Strapi: {len(cat_subs)} (visible after filter: {len(visible_subs)})")
    print(f"  Services in Strapi: {len(cat_svcs)} (visible after filter: {len(visible_svcs)})")

    if cat_hidden:
        print(f"  MENU STATUS: **NOT SHOWN** — category in hide-docs-123.json")
    elif len(visible_subs) == 0:
        print(f"  MENU STATUS: **NOT SHOWN** — all subcategories hidden")
    else:
        print(f"  MENU STATUS: Partially/fully visible ({len(visible_subs)} subcategories)")

    print(f"\n  Strapi subcategories:")
    for sub in sorted(cat_subs, key=lambda x: x.get("slug", "")):
        sub_hidden = sub["slug"] in hide_sub
        sub_svcs = [s for s in cat_svcs if (s.get("subcategory") or {}).get("slug") == sub["slug"]]
        vis = [s for s in sub_svcs if s["slug"] not in hide_svc]
        menu = "HIDDEN" if (cat_hidden or sub_hidden) else f"visible ({len(vis)} services)"
        print(f"    - {sub['title']} [{sub['slug']}] -> {menu}")

    print(f"\n  Sample Excel subdiv2 (first 8):")
    for code in sorted(subdiv2_in_excel.keys())[:8]:
        print(f"    {code} {subdiv2_in_excel[code]}")
    if len(subdiv2_in_excel) > 8:
        print(f"    ... +{len(subdiv2_in_excel) - 8} more")

# Summary table
print("\n" + "=" * 80)
print("SUMMARY: WHY CLIENT SEES 1.1–1.5 AS MISSING")
print("=" * 80)
for sec, slug in EXCEL_TO_CAT_SLUG.items():
    c = next((x for x in india_cats if x.get("slug") == slug), None)
    if not c:
        print(f"{sec}: NOT IN STRAPI")
    elif slug in hide_cat:
        n_sub = len([s for s in subs if (s.get("category") or {}).get("slug") == slug])
        n_svc = len(
            [
                s
                for s in services
                if (s.get("subcategory") or {}).get("slug")
                in {x["slug"] for x in subs if (x.get("category") or {}).get("slug") == slug}
            ]
        )
        print(f"{sec} {c['title']}: IN STRAPI ({n_sub} subs, {n_svc} services) BUT HIDDEN BY FRONTEND FILTER")
    else:
        print(f"{sec} {c['title']}: IN STRAPI AND NOT CATEGORY-HIDDEN")

# What India Practice shows instead
print("\n" + "=" * 80)
print("WHAT SHOWS IN MENU TODAY (India Practice — not 1.1-1.5)")
print("=" * 80)
for c in sorted(india_cats, key=lambda x: x.get("slug", "")):
    hidden = c["slug"] in hide_cat
    if not hidden:
        print(f"  VISIBLE: {c['title']} ({c['slug']})")

# Word doc mapping
print("\n" + "=" * 80)
print("WORD DOC COVERAGE FOR 1.1–1.5")
print("=" * 80)
doc_map = {
    "1.1": "Content RA Assosciates 1.docx (GST + Customs; also contains 1.2 start)",
    "1.2": "Content RA Assosciates 1.docx (Direct Tax — same file as 1.1)",
    "1.3": "Content RA Assosciates 2.docx + 3.docx (Audit & Assurance)",
    "1.4": "Content RA Assosciates 1.docx / 3.docx (Accounting — verify import)",
    "1.5": "Content RA Assosciates 3.docx / 4.docx (Corporate Advisory)",
}
for sec, doc in doc_map.items():
    print(f"  {sec}: {doc}")

# Slug mismatch note
print("\n" + "=" * 80)
print("SLUG MISMATCH NOTE (hide list vs current Strapi slugs)")
print("=" * 80)
strapi_gst_slugs = [
    s["slug"]
    for s in services
    if (s.get("subcategory") or {}).get("slug") == "gst-services"
]
old_hide_slugs = ["registration-service", "return-filing-service", "gst-advisory-service", "litigation-support-services", "gst-refunds"]
print("Strapi GST service slugs:", strapi_gst_slugs)
print("hide-docs-123 service slugs (old):", old_hide_slugs)
print("Service-level hide list may be OUTDATED — but category/subcategory hide still blocks entire 1.1–1.5 branches")
